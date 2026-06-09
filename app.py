from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from flask_socketio import SocketIO, emit
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import json
import re
import base64
import uuid
import requests
from datetime import datetime
import threading

app = Flask(__name__)
app.config['SECRET_KEY'] = 'attendance_secret_key_2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DOWNLOAD_FOLDER'] = 'downloads'
socketio = SocketIO(app, cors_allowed_origins="*")

# =====================================================================
#  뿌리오 API 설정 — 실제 값으로 교체하세요
# =====================================================================
PPURIO_ACCOUNT  = 'nvr_4277003079'   # 뿌리오 계정 아이디
PPURIO_API_KEY  = 'fc4839e53562e9711a3a5b0a49894547a39b7ca087fe5e3af3016c726ea155f5'          # 뿌리오 연동 개발 인증키
PPURIO_SENDER   = '01047722491'           # 발신자 번호 (숫자만, 등록된 번호)
# =====================================================================

PPURIO_TOKEN_URL   = 'https://message.ppurio.com/v1/token'
PPURIO_MESSAGE_URL = 'https://message.ppurio.com/v1/message'

# In-memory storage
attendance_data = []
data_lock = threading.Lock()

COLUMNS = ['사번', '이름', '소속', '직책', '이메일', '연락처', '출석체크']


def normalize_phone(raw: str) -> str:
    """
    연락처를 숫자만 남기되, 원본이 010으로 시작하면 반드시 010을 유지한다.
    예) '010-1234-5678' → '01012345678'
         '1012345678'   → '1012345678'  (010 없으면 그대로)
    """
    raw = str(raw).strip()
    digits_only = re.sub(r'\D', '', raw)
    # 원본(하이픈 제거 전)이 010으로 시작하면 강제 보정
    stripped_raw = re.sub(r'[\s\-]', '', raw)
    if stripped_raw.startswith('010') and not digits_only.startswith('010'):
        digits_only = '010' + digits_only
    return digits_only


def get_ppurio_token() -> str:
    """뿌리오 액세스 토큰 발급"""
    credential = base64.b64encode(f'{PPURIO_ACCOUNT}:{PPURIO_API_KEY}'.encode()).decode()
    
    session = requests.Session()
adapter = requests.adapters.HTTPAdapter(max_retries=3)
session.mount("https://", adapter)
    
        
    resp = session.post(
        PPURIO_TOKEN_URL,
        headers={'Authorization': f'Basic {credential}', 'Content-Type': 'application/json'},
        timeout=30
    )
    resp.raise_for_status()
    data = resp.json()
    return data['token']


def send_sms_via_ppurio(targets: list, message: str) -> dict:
    """
    targets: [{'to': '01012345678', 'name': '홍길동'}, ...]
    message: 발송 내용
    """
    token = get_ppurio_token()
    msg_type = 'LMS' if len(message.encode('euc-kr', errors='replace')) > 90 else 'SMS'

    payload = {
        'account':       PPURIO_ACCOUNT,
        'messageType':   msg_type,
        'content':       message,
        'from':          PPURIO_SENDER,
        'duplicateFlag': 'Y',
        'targetCount':   len(targets),
        'targets':       targets,
        'refKey':        uuid.uuid4().hex[:32],
    }

    resp = requests.post(
        PPURIO_MESSAGE_URL,
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type':  'application/json',
        },
        json=payload,
        timeout=15
    )
    return resp.json()


def load_excel(filepath):
    global attendance_data
    df = pd.read_excel(filepath, dtype=str)
    df.columns = df.columns.str.strip()
    for col in COLUMNS[:-1]:
        if col not in df.columns:
            df[col] = ''
    df['출석체크'] = ''
    with data_lock:
        attendance_data = df[COLUMNS].fillna('').to_dict('records')


def save_excel_to_download():
    os.makedirs(app.config['DOWNLOAD_FOLDER'], exist_ok=True)
    filepath = os.path.join(app.config['DOWNLOAD_FOLDER'], 'attendance_result.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '출석부'

    header_fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font  = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
    )

    col_widths = [12, 10, 15, 12, 25, 16, 20]
    for i, (col, width) in enumerate(zip(COLUMNS, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, header_align, border
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30

    attended_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    normal_fill   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    alt_fill      = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    cell_font  = Font(name='맑은 고딕', size=10)
    cell_align = Alignment(horizontal='center', vertical='center')

    with data_lock:
        rows = list(attendance_data)

    for r_idx, row in enumerate(rows, 2):
        is_attended = row.get('출석체크', '') != ''
        base_fill = attended_fill if is_attended else (normal_fill if r_idx % 2 == 0 else alt_fill)
        for c_idx, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(col, ''))
            cell.fill, cell.font, cell.alignment, cell.border = base_fill, cell_font, cell_align, border
        ws.row_dimensions[r_idx].height = 22

    wb.save(filepath)
    return filepath


# ── Routes ──────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('admin'))

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/user')
def user():
    return render_template('user.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'})
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'attendance.xlsx')
    file.save(filepath)
    try:
        load_excel(filepath)
        with data_lock:
            data = list(attendance_data)
        socketio.emit('data_updated', {'data': data})
        return jsonify({'success': True, 'message': f'{len(data)}명의 데이터가 로드되었습니다.', 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': f'파일 처리 오류: {str(e)}'})

@app.route('/api/data')
def get_data():
    with data_lock:
        data = list(attendance_data)
    return jsonify({'data': data})

@app.route('/api/attend', methods=['POST'])
def mark_attendance():
    body    = request.get_json()
    emp_id  = str(body.get('emp_id', '')).strip()
    name    = str(body.get('name',   '')).strip()
    if not emp_id or not name:
        return jsonify({'success': False, 'message': '사번과 이름을 입력해주세요.'})
    with data_lock:
        if not attendance_data:
            return jsonify({'success': False, 'message': '관리자가 아직 명단을 업로드하지 않았습니다.'})
        matched = next((r for r in attendance_data
                        if str(r.get('사번','')).strip() == emp_id
                        and str(r.get('이름','')).strip() == name), None)
        if not matched:
            return jsonify({'success': False, 'message': '일치하는 사번/이름 정보가 없습니다.'})
        if matched.get('출석체크'):
            return jsonify({'success': False, 'message': f'이미 출석 처리되었습니다. ({matched["출석체크"]})'})
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        matched['출석체크'] = timestamp
        data = list(attendance_data)
    socketio.emit('data_updated', {'data': data})
    return jsonify({'success': True, 'message': f'{name}님 출석이 확인되었습니다!', 'time': timestamp})

@app.route('/api/download')
def download_file():
    with data_lock:
        if not attendance_data:
            return jsonify({'success': False, 'message': '다운로드할 데이터가 없습니다.'})
    try:
        filepath = save_excel_to_download()
        return send_file(filepath, as_attachment=True,
                         download_name=f'출석부_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/reset', methods=['POST'])
def reset_data():
    global attendance_data
    with data_lock:
        attendance_data = []
    socketio.emit('data_updated', {'data': []})
    return jsonify({'success': True})

# ── SMS API ─────────────────────────────────────────────────────────

@app.route('/api/sms/preview', methods=['POST'])
def sms_preview():
    """발송 대상 미리보기 (실제 발송 없음)"""
    body    = request.get_json()
    mode    = body.get('mode', 'all')   # 'all' | 'absent'
    message = body.get('message', '').strip()

    if not message:
        return jsonify({'success': False, 'message': '메시지를 입력해주세요.'})

    with data_lock:
        rows = list(attendance_data)

    if not rows:
        return jsonify({'success': False, 'message': '명단 데이터가 없습니다.'})

    targets = []
    skipped = []
    for r in rows:
        if mode == 'absent' and r.get('출석체크', '').strip():
            continue
        raw_phone = r.get('연락처', '')
        phone = normalize_phone(raw_phone)
        if not phone or not phone.startswith('0') or len(phone) < 10:
            skipped.append({'name': r.get('이름',''), 'phone': raw_phone, 'reason': '유효하지 않은 번호'})
            continue
        targets.append({'name': r.get('이름',''), 'phone': phone, 'dept': r.get('소속','')})

    return jsonify({
        'success':  True,
        'count':    len(targets),
        'targets':  targets,
        'skipped':  skipped,
        'mode':     mode,
    })


@app.route('/api/sms/send', methods=['POST'])
def sms_send():
    """실제 뿌리오 API 호출하여 문자 발송"""
    body    = request.get_json()
    mode    = body.get('mode', 'all')
    message = body.get('message', '').strip()

    if not message:
        return jsonify({'success': False, 'message': '메시지를 입력해주세요.'})

    with data_lock:
        rows = list(attendance_data)

    if not rows:
        return jsonify({'success': False, 'message': '명단 데이터가 없습니다.'})

    # 발송 대상 구성
    ppurio_targets = []
    skipped = []
    for r in rows:
        if mode == 'absent' and r.get('출석체크', '').strip():
            continue
        raw_phone = r.get('연락처', '')
        phone = normalize_phone(raw_phone)
        if not phone or not phone.startswith('0') or len(phone) < 10:
            skipped.append(r.get('이름', ''))
            continue
        ppurio_targets.append({'to': phone, 'name': r.get('이름', '')})

    if not ppurio_targets:
        return jsonify({'success': False, 'message': '발송 가능한 유효한 번호가 없습니다.'})

    try:
        result = send_sms_via_ppurio(ppurio_targets, message)
        success = str(result.get('code', '')) == '1000'
        return jsonify({
            'success':    success,
            'message':    f'{"발송 완료" if success else "발송 실패"}: {result.get("description", "")}',
            'sent_count': len(ppurio_targets),
            'skipped':    skipped,
            'api_result': result,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'API 오류: {str(e)}'})


@socketio.on('connect')
def on_connect():
    with data_lock:
        data = list(attendance_data)
    emit('data_updated', {'data': data})


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    socketio.run(app, debug=False, host='0.0.0.0', port=port)